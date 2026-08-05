"""
Chargeur de fichiers XLSX pour le moteur d'intelligence documentaire GalSen IA.
"""

import os
from .document_loader_factory import BaseDocumentLoader
from .types import DocumentItem, DocumentType
import logging

logger = logging.getLogger(__name__)


class XLSXLoader(BaseDocumentLoader):
    """Chargeur pour les fichiers XLSX (.xlsx)."""

    # Extensions supportées
    supported_extensions = ['.xlsx']

    def load(self, source: str, **kwargs) -> DocumentItem:
        """
        Charge un fichier XLSX et retourne un DocumentItem.

        Args:
            source: Chemin vers le fichier XLSX
            **kwargs: Options supplémentaires

        Returns:
            DocumentItem représentant le fichier chargé
        """
        if not os.path.isfile(source):
            raise FileNotFoundError(f"Fichier non trouvé: {source}")

        try:
            import openpyxl
        except ImportError:
            message = (
                "openpyxl est requis pour charger les fichiers XLSX. "
                "Installez-le avec: pip install openpyxl"
            )
            logger.error(message)
            raise ImportError(message)

        # Charger le classeur
        workbook = openpyxl.load_workbook(source, data_only=True)

        # Extraire les données de toutes les feuilles
        text_content = []
        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            text_content.append(f"Feuille: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                # Convertir chaque cellule en chaîne et ignorer les None
                row_str = '\t'.join(['' if cell is None else str(cell) for cell in row])
                if row_str.strip():  # Ignorer les lignes complètement vides
                    text_content.append(row_str)
            text_content.append("")  # Ligne vide entre les feuilles

        full_text = '\n'.join(text_content)

        # Extraire les métadonnées du classeur
        props = workbook.properties
        metadata = {
            "size": os.path.getsize(source),
            "created": os.path.getctime(source),
            "modified": os.path.getmtime(source),
            "source": os.path.abspath(source),
            "filename": os.path.basename(source),
            "extension": ".xlsx",
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names
        }

        if props.creator:
            metadata["creator"] = props.creator
        if props.created:
            metadata["created"] = props.created.timestamp()
        if props.modified:
            metadata["modified"] = props.modified.timestamp()
        if props.title:
            metadata["title"] = props.title
        if props.description:
            metadata["description"] = props.description
        if props.keywords:
            # Les mots-clés peuvent être une chaîne séparée par des virgules
            if isinstance(props.keywords, str):
                metadata["keywords"] = [k.strip() for k in props.keywords.split(',')]
            else:
                metadata["keywords"] = list(props.keywords)
        if props.category:
            metadata["category"] = props.category

        # Utiliser le titre du classeur ou le nom du fichier
        title = metadata.get("title") or os.path.splitext(os.path.basename(source))[0]

        # Créer l'élément de document
        document = DocumentItem(
            document_id="",
            document_type=DocumentType.XLSX,
            title=title,
            content=full_text,
            metadata=metadata
        )

        return document

    def supports_type(self, document_type: str) -> bool:
        """Vérifie si le chargeur supporte le type de document donné."""
        return document_type in [
            DocumentType.XLSX.value,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        ]